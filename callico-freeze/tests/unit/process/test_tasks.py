import logging
import uuid
from datetime import timedelta

import pytest
from openpyxl import load_workbook

from callico.annotations.models import Annotation, Task, TaskState, TaskUser
from callico.process.tasks import csv_export, xlsx_export
from callico.projects.models import CSV_SUPPORTED_CAMPAIGN_MODES, XLSX_SUPPORTED_CAMPAIGN_MODES, CampaignMode, Element

pytestmark = pytest.mark.django_db


@pytest.mark.parametrize("mode", [mode for mode in CampaignMode if mode not in CSV_SUPPORTED_CAMPAIGN_MODES])
def test_csv_export_wrong_campaign_mode(mode, campaign, process_with_celery_mock):
    campaign.mode = mode
    campaign.save()

    with pytest.raises(AssertionError, match="CSV export for this campaign mode is not yet supported"):
        csv_export(**{"campaign_id": str(campaign.id)})

    campaign.refresh_from_db()
    assert not campaign.csv_export


@pytest.mark.parametrize("mode", CSV_SUPPORTED_CAMPAIGN_MODES)
def test_csv_export_annotation_value_error(caplog, mode, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
        ]
    }
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    user_task = task.user_tasks.create(user=contributor.user, state=TaskState.Annotated)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=20),
        # Invalid value
        value={},
    )

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.csv_export
    assert len(caplog.record_tuples) == 1
    _module, level, message = caplog.record_tuples[0]
    assert level == logging.ERROR
    assert message.startswith(f"Failed to export the last annotation on user task {user_task.id} in the CSV")


@pytest.mark.parametrize("mode", CSV_SUPPORTED_CAMPAIGN_MODES)
def test_csv_export_unassigned_task(mode, managed_campaign, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 0

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.csv_export


@pytest.mark.parametrize("mode", CSV_SUPPORTED_CAMPAIGN_MODES)
@pytest.mark.parametrize(
    "state", [state for state in TaskState if state not in [TaskState.Annotated, TaskState.Validated]]
)
def test_csv_export_wrong_user_task_state(mode, state, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    task = managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    task.user_tasks.create(user=contributor.user, state=state)
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 1

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.csv_export


@pytest.mark.parametrize("mode", CSV_SUPPORTED_CAMPAIGN_MODES)
def test_csv_export_ignore_preview_task(mode, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    task = managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    task.user_tasks.create(user=contributor.user, state=TaskState.Annotated, is_preview=True)
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 1

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.csv_export


@pytest.mark.parametrize("mode", CSV_SUPPORTED_CAMPAIGN_MODES)
def test_csv_export_multiple_user_tasks(mode, managed_campaign, contributor, new_contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
        ]
    }
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    TaskUser.objects.bulk_create(
        [TaskUser(task=task, user=user, state=TaskState.Pending) for user in [contributor.user, new_contributor]]
    )
    dog = project.classes.create(name="dog", provider=project.provider, provider_object_id=str(uuid.uuid4()))
    line = project.types.get(name="Line")
    for user_task in task.user_tasks.all():
        # Building a generic Annotation that works for all modes
        Annotation.objects.create(
            user_task=user_task,
            duration=timedelta(seconds=20),
            value={
                "transcription": {str(element.id): {"text": f"An annotation for the element {element.id}"}},
                "values": [
                    {
                        "value": "Harry",
                        "entity_type": "first_name",
                        "instruction": "The first name",
                    }
                ],
                "classification": str(dog.id),
                "entities": [{"entity_type": "person", "offset": 0, "length": 5}],
                "groups": [
                    {"elements": ["11111111-1111-1111-1111-111111111111", "22222222-2222-2222-2222-222222222222"]}
                ],
                "elements": [{"polygon": [[0, 0], [100, 0], [100, 50], [0, 50], [0, 0]], "element_type": str(line.id)}],
            },
        )
        user_task.state = TaskState.Annotated
        user_task.save()

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 3
    assert lines[1] != lines[2]


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_transcription_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Transcription
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()

    # Create element children
    transcriptions = ["This is an annotation", "on several", "children"]
    line_type = element.project.types.get(name="Line")
    children = Element.objects.bulk_create(
        Element(
            name=f"Line {i}",
            type=line_type,
            parent=element,
            project=element.project,
            provider=element.provider,
            provider_object_id=str(uuid.uuid4()),
            image=element.image,
            polygon=[[1, 2], [2, 3], [3, 4]],
            order=i,
        )
        for i, transcription in enumerate(transcriptions)
    )

    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "transcription": {
                str(elt.id): {"text": transcription}
                for elt, transcription in zip(
                    children + [element], transcriptions + [f"An annotation for the element {element.id}"]
                )
            },
        },
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.read().decode("utf8").split("\r\n")[:-1]
    assert len(lines) == 2
    assert (
        lines[0]
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,transcriptions"
    )
    assert (
        lines[1]
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}","An annotation for the element {element.id}\nThis is an annotation\non several\nchildren"'
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_entity_form_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.EntityForm
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
            {"entity_type": "last_name", "instruction": "The last name"},
            {"entity_type": "city", "instruction": "The location"},
            {
                "mode": "group",
                "legend": "Author",
                "fields": [
                    {"entity_type": "author_first_name", "instruction": "The first name"},
                ],
            },
        ]
    }
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "values": [
                {
                    "value": "Harry",
                    "entity_type": "first_name",
                    "instruction": "The first name",
                },
                {
                    "value": "Potter",
                    "entity_type": "last_name",
                    "instruction": "The last name",
                },
                {
                    "value": "Little Whinging",
                    "entity_type": "city",
                    "instruction": "The location",
                },
                {
                    "value": "Luna",
                    "entity_type": "author_first_name",
                    "instruction": "The first name",
                },
            ]
        },
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 2
    assert (
        lines[0].decode("utf8")
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,The first name (first_name),The last name (last_name),The location (city),Author > The first name (author_first_name)\r\n"
    )
    assert (
        lines[1].decode("utf8")
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}",Harry,Potter,Little Whinging,Luna\r\n'
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_classification_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Classification
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    dog = project.classes.create(name="dog", provider=project.provider, provider_object_id=str(uuid.uuid4()))
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={"classification": str(dog.id)},
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 2
    assert (
        lines[0].decode("utf8")
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,ml_class_callico_id,ml_class_provider_id,ml_class_name\r\n"
    )
    assert (
        lines[1].decode("utf8")
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}",{dog.id},{dog.provider_object_id},dog\r\n'
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_entity_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Entity
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "entities": [
                {"entity_type": "person", "offset": 0, "length": 5},
                {"entity_type": "person", "offset": 6, "length": 6},
            ]
        },
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 2
    assert (
        lines[0].decode("utf8")
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,number_of_annotated_entities\r\n"
    )
    assert (
        lines[1].decode("utf8")
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}",2\r\n'
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_element_group_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.ElementGroup
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "groups": [
                {"elements": ["11111111-1111-1111-1111-111111111111"]},
                {"elements": ["22222222-2222-2222-2222-222222222222"]},
            ]
        },
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 2
    assert (
        lines[0].decode("utf8")
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,number_of_annotated_groups\r\n"
    )
    assert (
        lines[1].decode("utf8")
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}",2\r\n'
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_csv_export_elements_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Elements
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    line = project.types.get(name="Line")
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "elements": [
                {"polygon": [[0, 0], [100, 0], [100, 50], [0, 50], [0, 0]], "element_type": str(line.id)},
                {"polygon": [[100, 100], [200, 100], [200, 150], [100, 150], [100, 100]], "element_type": str(line.id)},
            ],
        },
    )

    csv_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.csv_export
    assert managed_campaign.csv_export.name == f"csv_exports/export-{str(managed_campaign.id)[:8]}.csv"
    lines = managed_campaign.csv_export.readlines()
    assert len(lines) == 2
    assert (
        lines[0].decode("utf8")
        == "id,state,annotator_email,created,completion_time_in_seconds,number_of_comments,callico_task_url,provider_element_url,iiif_url,element_thumbnail_url,number_of_annotated_elements\r\n"
    )
    assert (
        lines[1].decode("utf8")
        == f'{user_task.id},{user_task.get_state_display()},contributor@callico.org,{user_task.created},{str(duration or "")},1,https://callico.test{user_task.annotate_url},{element.provider_url},{element.image.iiif_url},"{element.build_thumbnail(size_max_height=1600)}",2\r\n'
    )


@pytest.mark.parametrize("mode", [mode for mode in CampaignMode if mode not in XLSX_SUPPORTED_CAMPAIGN_MODES])
def test_xlsx_export_wrong_campaign_mode(mode, campaign, process_with_celery_mock):
    campaign.mode = mode
    campaign.save()

    with pytest.raises(AssertionError, match="XLSX export for this campaign mode is not yet supported"):
        xlsx_export(**{"campaign_id": str(campaign.id)})

    campaign.refresh_from_db()
    assert not campaign.xlsx_export


@pytest.mark.parametrize("mode", XLSX_SUPPORTED_CAMPAIGN_MODES)
def test_xlsx_export_annotation_value_error(caplog, mode, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
        ]
    }
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    user_task = task.user_tasks.create(user=contributor.user, state=TaskState.Annotated)
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=20),
        # Invalid value
        value={},
    )

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.xlsx_export
    assert len(caplog.record_tuples) == 1
    _module, level, message = caplog.record_tuples[0]
    assert level == logging.ERROR
    assert message.startswith(f"Failed to export the last annotation on user task {user_task.id} in the XLSX")


@pytest.mark.parametrize("mode", XLSX_SUPPORTED_CAMPAIGN_MODES)
def test_xlsx_export_unassigned_task(mode, managed_campaign, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 0

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.xlsx_export


@pytest.mark.parametrize("mode", XLSX_SUPPORTED_CAMPAIGN_MODES)
@pytest.mark.parametrize(
    "state", [state for state in TaskState if state not in [TaskState.Annotated, TaskState.Validated]]
)
def test_xlsx_export_wrong_user_task_state(mode, state, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    task = managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    task.user_tasks.create(user=contributor.user, state=state)
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 1

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.xlsx_export


@pytest.mark.parametrize("mode", XLSX_SUPPORTED_CAMPAIGN_MODES)
def test_xlsx_export_ignore_preview_task(mode, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = mode
    managed_campaign.save()
    task = managed_campaign.tasks.create(element=managed_campaign.project.elements.filter(image__isnull=False).first())
    task.user_tasks.create(user=contributor.user, state=TaskState.Annotated, is_preview=True)
    assert Task.objects.filter(campaign=managed_campaign).count() == 1
    assert TaskUser.objects.filter(task__campaign=managed_campaign).count() == 1

    with pytest.raises(
        Exception, match="No valid results to be exported were found for this campaign, no file will be created"
    ):
        xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert not managed_campaign.xlsx_export


@pytest.mark.parametrize("mode", XLSX_SUPPORTED_CAMPAIGN_MODES)
def test_xlsx_export_multiple_user_tasks(
    mode, managed_campaign, contributor, new_contributor, process_with_celery_mock
):
    managed_campaign.mode = mode
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
        ]
    }
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    TaskUser.objects.bulk_create(
        [TaskUser(task=task, user=user, state=TaskState.Pending) for user in [contributor.user, new_contributor]]
    )
    dog = project.classes.create(name="dog", provider=project.provider, provider_object_id=str(uuid.uuid4()))
    line = project.types.get(name="Line")
    for user_task in task.user_tasks.all():
        # Building a generic Annotation that works for all modes
        Annotation.objects.create(
            user_task=user_task,
            duration=timedelta(seconds=20),
            value={
                "transcription": {str(element.id): {"text": f"An annotation for the element {element.id}"}},
                "values": [
                    {
                        "value": "Harry",
                        "entity_type": "first_name",
                        "instruction": "The first name",
                    }
                ],
                "classification": str(dog.id),
                "entities": [{"entity_type": "person", "offset": 0, "length": 5}],
                "groups": [
                    {"elements": ["11111111-1111-1111-1111-111111111111", "22222222-2222-2222-2222-222222222222"]}
                ],
                "elements": [{"polygon": [[0, 0], [100, 0], [100, 50], [0, 50], [0, 0]], "element_type": str(line.id)}],
            },
        )
        user_task.state = TaskState.Annotated
        user_task.save()

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 3
    assert rows[1] != rows[2]


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_transcription_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Transcription
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()

    # Create element children
    transcriptions = ["This is an annotation", "on several", "children"]
    line_type = element.project.types.get(name="Line")
    children = Element.objects.bulk_create(
        Element(
            name=f"Line {i}",
            type=line_type,
            parent=element,
            project=element.project,
            provider=element.provider,
            provider_object_id=str(uuid.uuid4()),
            image=element.image,
            polygon=[[1, 2], [2, 3], [3, 4]],
            order=i,
        )
        for i, transcription in enumerate(transcriptions)
    )

    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "transcription": {
                str(elt.id): {"text": transcription}
                for elt, transcription in zip(
                    children + [element], transcriptions + [f"An annotation for the element {element.id}"]
                )
            },
        },
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "transcriptions",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        f"An annotation for the element {element.id}\nThis is an annotation\non several\nchildren",
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_entity_form_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.EntityForm
    managed_campaign.configuration = {
        "fields": [
            {"entity_type": "first_name", "instruction": "The first name"},
            {"entity_type": "last_name", "instruction": "The last name"},
            {"entity_type": "city", "instruction": "The location"},
            {
                "mode": "group",
                "legend": "Author",
                "fields": [
                    {"entity_type": "author_first_name", "instruction": "The first name"},
                ],
            },
        ]
    }
    managed_campaign.save()
    element = managed_campaign.project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "values": [
                {
                    "value": "Harry",
                    "entity_type": "first_name",
                    "instruction": "The first name",
                },
                {
                    "value": "Potter",
                    "entity_type": "last_name",
                    "instruction": "The last name",
                },
                {
                    "value": "Little Whinging",
                    "entity_type": "city",
                    "instruction": "The location",
                },
                {
                    "value": "Luna",
                    "entity_type": "author_first_name",
                    "instruction": "The first name",
                },
            ]
        },
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "The first name (first_name)",
        "The last name (last_name)",
        "The location (city)",
        "Author > The first name (author_first_name)",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        "Harry",
        "Potter",
        "Little Whinging",
        "Luna",
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_classification_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Classification
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    dog = project.classes.create(name="dog", provider=project.provider, provider_object_id=str(uuid.uuid4()))
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={"classification": str(dog.id)},
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "ml_class_callico_id",
        "ml_class_provider_id",
        "ml_class_name",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        str(dog.id),
        dog.provider_object_id,
        "dog",
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_entity_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Entity
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "entities": [
                {"entity_type": "person", "offset": 0, "length": 5},
                {"entity_type": "person", "offset": 6, "length": 6},
            ]
        },
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "number_of_annotated_entities",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        2,
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_element_group_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.ElementGroup
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "groups": [
                {"elements": ["11111111-1111-1111-1111-111111111111"]},
                {"elements": ["22222222-2222-2222-2222-222222222222"]},
            ]
        },
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "number_of_annotated_groups",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        2,
    )


@pytest.mark.parametrize("state", [TaskState.Annotated, TaskState.Validated])
@pytest.mark.parametrize("duration", [None, 20])
def test_xlsx_export_elements_campaign(state, duration, managed_campaign, contributor, process_with_celery_mock):
    managed_campaign.mode = CampaignMode.Elements
    managed_campaign.save()
    project = managed_campaign.project
    element = project.elements.filter(image__isnull=False).first()
    task = managed_campaign.tasks.create(element=element)
    task.comments.create(user=contributor.user, content="Oops")
    user_task = task.user_tasks.create(user=contributor.user, state=state)
    formatted_created_date = user_task.created.replace(tzinfo=None).replace(
        microsecond=round(user_task.created.microsecond / 1000) * 1000
    )
    line = project.types.get(name="Line")
    Annotation.objects.create(
        user_task=user_task,
        duration=timedelta(seconds=duration) if duration else None,
        value={
            "elements": [
                {"polygon": [[0, 0], [100, 0], [100, 50], [0, 50], [0, 0]], "element_type": str(line.id)},
                {"polygon": [[100, 100], [200, 100], [200, 150], [100, 150], [100, 100]], "element_type": str(line.id)},
            ],
        },
    )

    xlsx_export(**{"campaign_id": str(managed_campaign.id)})

    managed_campaign.refresh_from_db()
    assert managed_campaign.xlsx_export
    assert managed_campaign.xlsx_export.name == f"xlsx_exports/export-{str(managed_campaign.id)[:8]}.xlsx"
    worksheet = load_workbook(filename=managed_campaign.xlsx_export.path).active
    rows = list(worksheet.values)
    assert len(rows) == 2
    assert rows[0] == (
        "id",
        "state",
        "annotator_email",
        "created",
        "completion_time_in_seconds",
        "number_of_comments",
        "callico_task_url",
        "provider_element_url",
        "iiif_url",
        "element_thumbnail_url",
        "number_of_annotated_elements",
    )
    assert rows[1] == (
        str(user_task.id),
        user_task.get_state_display(),
        "contributor@callico.org",
        formatted_created_date,
        duration,
        1,
        f"https://callico.test{user_task.annotate_url}",
        element.provider_url,
        element.image.iiif_url,
        element.build_thumbnail(size_max_height=1600),
        2,
    )
